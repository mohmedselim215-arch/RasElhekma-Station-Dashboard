import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_dashboard_sheet_with_link():
    excel_path = r"D:\log\RAAS ELHEKMA STATION LOG.xlsm"
    
    # التأكد من وجود الملف أولاً
    if not os.path.exists(excel_path):
        print(f"⚠️ لم يتم العثور على ملف الإكسيل في المسار المحدد: {excel_path}")
        return

    try:
        print("⏳ جاري فتح ملف الإكسيل وقراءة البيانات...")
        # التحميل مع الحفاظ على أكواد الماكرو الـ VBA بالداخل (keep_vba=True)
        wb = openpyxl.load_workbook(excel_path, keep_vba=True)
        
        # إذا كان الشيت موجوداً مسبقاً، نقوم بحذفه لإعادة بنائه بشكل نظيف
        if "داش بورد" in wb.sheetnames:
            del wb["داش بورد"]
            
        # إنشاء شيت جديد وتحديده ليكون الأول (Index = 0)
        ws = wb.create_sheet(title="داش بورد", index=0)
        
        # تفعيل إظهار خطوط الشبكة (Gridlines) ليكون المظهر مألوفاً
        ws.views.sheetView[0].showGridLines = True
        
        # 1. تصميم عنوان رئيسي جذاب داخل الشيت
        ws['B2'] = "📊 لوحة التحكم التفاعلية - محطة رأس الحكمة (S-19)"
        ws['B2'].font = Font(name='Segoe UI', size=16, bold=True, color='0F172A')
        
        ws['B4'] = "اضغط على الزر أدناه لفتح الداش بورد التفاعلية في المتصفح وتحديث البيانات تلقائياً:"
        ws['B4'].font = Font(name='Segoe UI', size=11, italic=True, color='475569')
        
        # 2. تصميم خلية الرابط كأنها "زر تفاعلي" (Styled Button)
        button_cell = ws['B6']
        button_cell.value = "🚀 تشغيل لوحة التحكم التفاعلية"
        button_cell.hyperlink = "http://localhost:8501"
        
        # تنسيق الزر (لون أزرق احترافي مريح للعين، خط سميك مسطر، محاذاة في المنتصف)
        button_cell.font = Font(name='Segoe UI', size=13, bold=True, color='FFFFFF', underline='single')
        button_cell.fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # كحلي غامق
        button_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ضبط مقاسات الأعمدة والصفوف لتظهر بشكل مريح جداً للعين
        ws.column_dimensions['B'].width = 35
        ws.row_dimensions[6].height = 40  # جعل الزر عريضاً وسهل الضغط
        
        # حفظ الملف بنفس اسمه وصيغته دون التأثير على محتوياته الأخرى
        print("💾 جاري حفظ التعديلات داخل ملف الإكسيل...")
        wb.save(excel_path)
        print("✅ تم بنجاح إنشاء شيت 'داش بورد' في أول الملف وبداخله الرابط التفاعلي!")
        
    except Exception as e:
        print(f"❌ حدث خطأ غير متوقع: {e}")
        print("تأكد من إغلاق ملف الإكسيل قبل تشغيل هذا الكود لتجنب حدوث خطأ Permission Denied.")

if __name__ == "__main__":
    create_dashboard_sheet_with_link()